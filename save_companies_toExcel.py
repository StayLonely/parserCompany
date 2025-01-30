import pandas as pd
from openpyxl import load_workbook

def save_companies_toEscel(all_companies_data):

    filename = 'company_data.xlsx'

    new_df = pd.DataFrame(all_companies_data)

    try:

        book = load_workbook(filename)
        sheet_name = book.sheetnames[0]
        existing_df = pd.read_excel(filename, sheet_name=sheet_name)


        combined_df = pd.concat([existing_df, new_df], ignore_index=True)


        with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
            combined_df.to_excel(writer, index=False, sheet_name=sheet_name)


            worksheet = writer.sheets[sheet_name]


            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

    except FileNotFoundError:

        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            new_df.to_excel(writer, index=False, sheet_name='Sheet1')


            worksheet = writer.sheets['Sheet1']

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"Данные о компаниях сохранены в файле {filename}")